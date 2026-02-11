import pandas as pd
from pathlib import Path
import openpyxl

class ExcelLoader:
    def __init__(self, excel_path):
        self.excel_path = Path(excel_path)
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.excel_path}")

    def load_text_sheet(self, sheet_name, required_columns=None):
        """Load text data sheet into list of dicts."""
        try:
            df = pd.read_excel(self.excel_path, sheet_name=sheet_name, dtype=str)
            df = df.fillna('') # Replace NaN with empty string
            
            # Validate columns
            if required_columns:
                missing = [col for col in required_columns if col not in df.columns]
                if missing:
                    raise ValueError(f"Missing required columns in sheet '{sheet_name}': {missing}")
            
            return df.to_dict('records')
        except Exception as e:
            raise ValueError(f"Error loading sheet '{sheet_name}': {str(e)}")

    def load_config_sheet(self, sheet_name):
        """Load table mapping config sheet."""
        required_cols = ['TableName', 'SourceSheet', 'Range', 'WordBookmark']
        try:
            df = pd.read_excel(self.excel_path, sheet_name=sheet_name, dtype=str)
            df = df.fillna('')
            
            missing = [col for col in required_cols if col not in df.columns]
            if missing:
                raise ValueError(f"Config sheet '{sheet_name}' missing columns: {missing}")
                
            return df.to_dict('records')
        except ValueError as e:
            # If sheet doesn't exist, return empty list (optional config)
            # or raise error depending on strictness. Let's return empty for now if missing.
            # But wait, read_excel raises ValueError if sheet not found.
            # If config sheet is defined in YAML, it MUST exist.
            raise e
        except Exception as e:
            raise ValueError(f"Error loading config sheet '{sheet_name}': {str(e)}")

    def load_table_range(self, sheet_name, range_str):
        """Load specific Excel range as DataFrame."""
        try:
            # range_str format: "A1:D20" or just "A1" (to end)
            # using openpyxl engine via pandas is tricky for exact ranges
            # Better to use openpyxl directly to parse range or use pandas 'usecols' + 'skiprows'
            # But 'usecols' with 'A:D' and 'nrows' is cleaner.
            
            # Simple approach: Read whole sheet, then slice.
            # For large sheets this is inefficient, but for Phase 1 it's fine.
            # Optimization: Use openpyxl to get dimensions if needed.
            
            # Let's try parsing "A1:D20"
            start_cell, end_cell = range_str.split(':')
            
            # This is complex to do robustly with just pandas.
            # Standard approach:
            # 1. Load sheet with header=None (to get absolute indexing)
            # 2. Slice DataFrame using parse_range logic
            # 3. Promote first row to header
            
            # Actually, simpler: Use 'openpyxl' to read values from range
            wb = openpyxl.load_workbook(self.excel_path, data_only=True)
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found")
            
            ws = wb[sheet_name]
            data = ws[range_str] # variable_name["A1:C4"] returns tuple of rows
            
            # Convert to list of lists
            rows = []
            for row in data:
                rows.append([cell.value for cell in row])
                
            if not rows:
                return pd.DataFrame()
                
            # Assume first row of range is header
            headers = rows[0]
            data_rows = rows[1:]
            
            df = pd.DataFrame(data_rows, columns=headers)
            df = df.fillna('')
            return df

        except Exception as e:
            raise ValueError(f"Error loading table range '{range_str}' from '{sheet_name}': {str(e)}")

